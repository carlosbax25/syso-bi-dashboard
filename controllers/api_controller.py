from flask import Blueprint, request, jsonify, send_file
from datetime import date
from io import BytesIO

from services.dashboard_service import DashboardService

api_bp = Blueprint('api', __name__, url_prefix='/api')


def init_api(service: DashboardService):
    """Inicializa el blueprint API con la instancia del servicio."""

    @api_bp.route('/datos')
    def obtener_datos():
        """Endpoint principal: retorna órdenes filtradas, KPIs y datos de gráficos."""
        fecha_inicio = _parse_date(request.args.get('fecha_inicio'))
        fecha_fin = _parse_date(request.args.get('fecha_fin'))
        arls = request.args.getlist('arl') or None
        tipos = request.args.getlist('tipo_servicio') or None
        estados = request.args.getlist('estado') or None
        programas = request.args.getlist('programa') or None
        tareas = request.args.getlist('tarea') or None

        ordenes = service.obtener_ordenes_filtradas(
            fecha_inicio, fecha_fin, arls, tipos, estados, programas, tareas
        )
        kpis = service.calcular_kpis(ordenes)

        return jsonify({
            'ordenes': [o.to_dict() for o in ordenes],
            'kpis': kpis.to_dict(),
            'graficos': {
                'ordenes_por_arl': service.agrupar_ordenes_por_arl(ordenes),
                'ordenes_por_mes': service.agrupar_ordenes_por_mes(ordenes),
                'ordenes_por_servicio': service.agrupar_ordenes_por_servicio(ordenes),
                'ingresos_por_arl': service.agrupar_ingresos_por_arl(ordenes),
            },
            'pendientes_por_arl': service.analizar_pendientes_por_arl(ordenes),
            'cartera_por_arl': service.analizar_cartera(ordenes),
            'tiempos_proceso': service.analizar_tiempos_proceso(ordenes),
            'proyeccion_ingresos': service.proyectar_ingresos_por_arl(ordenes, meses_hist_limit=12),
            'filtros_disponibles': {
                'arls': service.obtener_arls(),
                'tipos_servicio': sorted(set(o.tipo_servicio for o in ordenes)),
                'estados': ['Recibida', 'Aceptada', 'Rechazada', 'Programada / Asignada',
                            'En Ejecución', 'Ejecutada', 'Soportes Radicados',
                            'Facturada', 'Reemplazada', 'Cancelada'],
                'programas': sorted(set(o.programa for o in ordenes)),
                'tareas': sorted(set(o.tarea for o in ordenes)),
            },
        })

    @api_bp.route('/filtros')
    def obtener_filtros():
        """Retorna opciones de filtros en cascada.
        Acepta query params para filtrar: arl, tipo_servicio, programa
        """
        arls_sel = request.args.getlist('arl') or None
        tipos_sel = request.args.getlist('tipo_servicio') or None
        programas_sel = request.args.getlist('programa') or None

        # Start with all orders
        ordenes = service._repo.obtener_ordenes()

        # Apply cascade: ARL filters everything below
        if arls_sel:
            arls_set = set(arls_sel)
            ordenes = [o for o in ordenes if o.arl in arls_set]

        tipos_disponibles = sorted(set(o.tipo_servicio for o in ordenes))

        # Tipo servicio filters programa and tarea
        if tipos_sel:
            tipos_set = set(tipos_sel)
            ordenes_tipo = [o for o in ordenes if o.tipo_servicio in tipos_set]
        else:
            ordenes_tipo = ordenes

        programas_disponibles = sorted(set(o.programa for o in ordenes_tipo))

        # Programa filters tarea
        if programas_sel:
            prog_set = set(programas_sel)
            ordenes_prog = [o for o in ordenes_tipo if o.programa in prog_set]
        else:
            ordenes_prog = ordenes_tipo

        tareas_disponibles = sorted(set(o.tarea for o in ordenes_prog))

        return jsonify({
            'arls': service.obtener_arls(),
            'tipos_servicio': tipos_disponibles,
            'programas': programas_disponibles,
            'tareas': tareas_disponibles,
        })

    @api_bp.route('/exportar')
    def exportar_excel():
        """Exporta las órdenes filtradas a Excel."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        fecha_inicio = _parse_date(request.args.get('fecha_inicio'))
        fecha_fin = _parse_date(request.args.get('fecha_fin'))
        arls = request.args.getlist('arl') or None
        tipos = request.args.getlist('tipo_servicio') or None
        estados = request.args.getlist('estado') or None
        programas = request.args.getlist('programa') or None
        tareas = request.args.getlist('tarea') or None

        ordenes = service.obtener_ordenes_filtradas(
            fecha_inicio, fecha_fin, arls, tipos, estados, programas, tareas
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ordenes de Servicio'

        headers = ['ID', 'Fecha', 'ARL', 'Empresa', 'Tipo Servicio', 'Programa', 'Tarea', 'Trabajadores', 'Estado', 'Valor Facturado']
        header_font = Font(bold=True, color='FFFFFF', size=10)
        header_fill = PatternFill(start_color='1F67AE', end_color='1F67AE', fill_type='solid')

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for i, o in enumerate(ordenes, 2):
            ws.cell(row=i, column=1, value=o.id)
            ws.cell(row=i, column=2, value=o.fecha.isoformat())
            ws.cell(row=i, column=3, value=o.arl)
            ws.cell(row=i, column=4, value=o.empresa)
            ws.cell(row=i, column=5, value=o.tipo_servicio)
            ws.cell(row=i, column=6, value=o.programa)
            ws.cell(row=i, column=7, value=o.tarea)
            ws.cell(row=i, column=8, value=o.cantidad_trabajadores)
            ws.cell(row=i, column=9, value=o.estado)
            ws.cell(row=i, column=10, value=o.valor_facturado)

        ws.freeze_panes = 'A2'

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Reporte_Ordenes_IPS_ARL.xlsx'
        )

    return api_bp


def _parse_date(value: str) -> date | None:
    """Parsea una fecha ISO. Retorna None si el valor es vacío o inválido."""
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None
